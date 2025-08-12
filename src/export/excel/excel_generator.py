"""
Excel Generator
Main engine for creating comprehensive Excel workbooks with financial analysis

Creates professional Excel workbooks with multiple worksheets containing
analysis results, charts, calculations, and formatted data suitable for
executive review and stakeholder sharing.
"""

import asyncio
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime
import tempfile

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.drawing.image import Image

from .data_formatting import ExcelFormatter
from .chart_embedding import ChartEmbedder
from .template_manager import ExcelTemplateManager


# Configure logging
logger = logging.getLogger(__name__)


class ExcelGenerator:
    """
    Professional Excel workbook generator for real estate analysis
    
    Creates comprehensive workbooks with multiple worksheets, embedded charts,
    and professional formatting suitable for executive presentation.
    """
    
    def __init__(self, temp_dir: Optional[Path] = None):
        """
        Initialize Excel generator
        
        Args:
            temp_dir: Directory for temporary files during generation
        """
        self.temp_dir = Path(temp_dir) if temp_dir else Path(tempfile.gettempdir()) / "excel_generation"
        self.temp_dir.mkdir(parents=True, exist_ok=True)
        
        # Initialize components
        self.formatter = ExcelFormatter()
        self.chart_embedder = ChartEmbedder()
        self.template_manager = ExcelTemplateManager()
        
        # Workbook components
        self.workbook: Optional[Workbook] = None
        self.worksheets: Dict[str, Worksheet] = {}
        
        logger.info(f"ExcelGenerator initialized with temp_dir: {self.temp_dir}")
    
    async def prepare_data(self, export_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Prepare data for Excel generation
        
        Args:
            export_data: Complete export data package
            
        Returns:
            Excel-specific data package with formatted tables and calculations
        """
        logger.info("Preparing data for Excel generation")
        
        excel_data = {
            # Core data from export package
            'analysis_results': export_data['analysis_results'],
            'ownership_flows': export_data['ownership_flows'],
            'rental_flows': export_data['rental_flows'],
            'session_data': export_data.get('session_data') or export_data.get('inputs', {}),
            'export_options': export_data.get('export_options', {}),
            
            # Excel-specific formatted data
            'formatted_tables': {},
            'summary_metrics': {},
            'chart_data': {},
            'worksheet_configs': {}
        }
        
        # Format data for different worksheets
        await self._prepare_summary_data(excel_data)
        await self._prepare_cash_flow_data(excel_data) 
        await self._prepare_calculations_data(excel_data)
        await self._prepare_assumptions_data(excel_data)
        
        logger.info("Excel data preparation completed")
        return excel_data
    
    async def render_charts(
        self, 
        export_data: Dict[str, Any], 
        resolution: int = 300
    ) -> Dict[str, Path]:
        """
        Render interactive charts as high-resolution images for Excel embedding
        
        Args:
            export_data: Export data with chart configurations
            resolution: Chart resolution in DPI
            
        Returns:
            Dictionary mapping chart names to image file paths
        """
        logger.info(f"Rendering charts for Excel at {resolution} DPI")
        
        chart_images = await self.chart_embedder.render_all_charts(
            export_data, resolution, self.temp_dir
        )
        
        logger.info(f"Rendered {len(chart_images)} charts for Excel")
        return chart_images
    
    async def generate_workbook(
        self,
        excel_data: Dict[str, Any],
        template_type: str = "detailed"
    ) -> Path:
        """
        Generate complete Excel workbook with all worksheets and formatting
        
        Args:
            excel_data: Prepared Excel data package
            template_type: Template type ("executive", "detailed", "investor")
            
        Returns:
            Path to generated Excel file
        """
        logger.info(f"Generating Excel workbook with template: {template_type}")
        
        try:
            # Initialize workbook
            self.workbook = Workbook()
            self.workbook.remove(self.workbook.active)  # Remove default sheet
            
            # Apply template configuration
            template_config = await self.template_manager.get_template_config(template_type)
            
            # Create worksheets based on template
            await self._create_worksheets(excel_data, template_config)
            
            # Apply professional styling
            await self._apply_workbook_styling(template_config)
            
            # Generate output file
            output_path = self.temp_dir / f"real_estate_analysis_{template_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # Save workbook
            self.workbook.save(output_path)
            
            logger.info(f"Excel workbook generated: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Excel workbook generation failed: {str(e)}")
            raise RuntimeError(f"Excel generation failed: {str(e)}") from e
    
    async def _prepare_summary_data(self, excel_data: Dict[str, Any]) -> None:
        """Prepare executive summary data"""
        analysis = excel_data['analysis_results']
        
        summary_data = {
            'recommendation': analysis.get('recommendation', 'UNKNOWN'),
            'confidence': analysis.get('confidence', 'Low'),
            'npv_difference': analysis.get('npv_difference', 0),
            'ownership_npv': analysis.get('ownership_npv', 0),
            'rental_npv': analysis.get('rental_npv', 0),
            'ownership_initial_investment': analysis.get('ownership_initial_investment', 0),
            'rental_initial_investment': analysis.get('rental_initial_investment', 0),
            'analysis_period': analysis.get('analysis_period', 25),
            'cost_of_capital': analysis.get('cost_of_capital', 8.0)
        }
        
        excel_data['summary_metrics'] = summary_data
    
    async def _prepare_cash_flow_data(self, excel_data: Dict[str, Any]) -> None:
        """Prepare cash flow analysis data"""
        ownership_flows = excel_data['ownership_flows']
        rental_flows = excel_data['rental_flows']
        
        # Normalize cash flow format for processing
        normalized_ownership = self._normalize_cash_flows(ownership_flows)
        normalized_rental = self._normalize_cash_flows(rental_flows)
        
        # Create formatted cash flow tables
        formatted_flows = {
            'ownership_table': await self.formatter.format_cash_flow_table(normalized_ownership, "ownership"),
            'rental_table': await self.formatter.format_cash_flow_table(normalized_rental, "rental"),
            'comparison_table': await self.formatter.create_comparison_table(normalized_ownership, normalized_rental)
        }
        
        excel_data['formatted_tables']['cash_flows'] = formatted_flows
    
    async def _prepare_calculations_data(self, excel_data: Dict[str, Any]) -> None:
        """Prepare detailed calculations data"""
        analysis = excel_data['analysis_results']
        
        # Extract calculation details
        calculations = {
            'npv_calculations': await self.formatter.format_npv_calculations(analysis),
            'mortgage_schedule': await self.formatter.format_mortgage_schedule(analysis),
            'tax_calculations': await self.formatter.format_tax_calculations(analysis),
            'terminal_value': await self.formatter.format_terminal_value(analysis)
        }
        
        excel_data['formatted_tables']['calculations'] = calculations
    
    async def _prepare_assumptions_data(self, excel_data: Dict[str, Any]) -> None:
        """Prepare input assumptions data"""
        session_data = excel_data['session_data']
        
        # Format assumptions by category
        assumptions = await self.formatter.format_assumptions_table(session_data)
        excel_data['formatted_tables']['assumptions'] = assumptions
    
    async def _create_worksheets(
        self, 
        excel_data: Dict[str, Any], 
        template_config: Dict[str, Any]
    ) -> None:
        """Create all worksheets based on template configuration"""
        
        worksheet_configs = template_config.get('worksheets', [])
        
        for config in worksheet_configs:
            worksheet_name = config['name']
            worksheet_type = config['type']
            
            # Create worksheet
            ws = self.workbook.create_sheet(title=worksheet_name)
            self.worksheets[worksheet_name] = ws
            
            # Populate worksheet based on type
            await self._populate_worksheet(ws, worksheet_type, excel_data, config)
    
    async def _populate_worksheet(
        self,
        ws: Worksheet,
        worksheet_type: str,
        excel_data: Dict[str, Any],
        config: Dict[str, Any]
    ) -> None:
        """Populate individual worksheet with data and formatting"""
        
        if worksheet_type == "executive_summary":
            await self._create_executive_summary_sheet(ws, excel_data, config)
        elif worksheet_type == "cash_flows":
            await self._create_cash_flows_sheet(ws, excel_data, config)
        elif worksheet_type == "charts":
            await self._create_charts_sheet(ws, excel_data, config)
        elif worksheet_type == "calculations":
            await self._create_calculations_sheet(ws, excel_data, config)
        elif worksheet_type == "assumptions":
            await self._create_assumptions_sheet(ws, excel_data, config)
        else:
            logger.warning(f"Unknown worksheet type: {worksheet_type}")
    
    async def _create_executive_summary_sheet(
        self,
        ws: Worksheet,
        excel_data: Dict[str, Any],
        config: Dict[str, Any]
    ) -> None:
        """Create executive summary worksheet"""
        summary = excel_data['summary_metrics']
        
        # Enhanced title section with professional styling
        ws['A1'] = "REAL ESTATE INVESTMENT ANALYSIS"
        ws['A1'].font = Font(name='Calibri', size=18, bold=True, color='FF6B6B')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')  # Merge cells for better title presentation
        
        ws['A2'] = f"Executive Summary & Strategic Recommendation - {datetime.now().strftime('%B %d, %Y')}"
        ws['A2'].font = Font(name='Calibri', size=12, italic=True, color='2D3436')
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A2:D2')
        
        # Add professional spacing
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 20
        
        # Enhanced recommendation section with better visual hierarchy
        row = 4
        ws[f'A{row}'] = "STRATEGIC RECOMMENDATION"
        ws[f'A{row}'].font = Font(name='Calibri', size=14, bold=True, color='2D3436')
        
        row += 1
        recommendation = summary['recommendation']
        confidence = summary.get('confidence', 'Medium')
        
        # Create a more professional recommendation display
        ws[f'A{row}'] = f"{recommendation.upper()}"
        ws[f'A{row}'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws[f'C{row}'] = f"Confidence: {confidence}"
        ws[f'C{row}'].font = Font(name='Calibri', size=11, italic=True, color='2D3436')
        ws[f'C{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Enhanced color coding with professional palette
        if recommendation == "BUY":
            ws[f'A{row}'].fill = PatternFill(start_color="00B894", end_color="00B894", fill_type="solid")
        elif recommendation == "RENT":
            ws[f'A{row}'].fill = PatternFill(start_color="FDCB6E", end_color="FDCB6E", fill_type="solid")
        
        # Add cell borders for professional appearance
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        ws[f'A{row}'].border = thin_border
        ws.row_dimensions[row].height = 25
        
        # Enhanced key metrics section with professional table formatting
        row += 3
        ws[f'A{row}'] = "KEY FINANCIAL METRICS"
        ws[f'A{row}'].font = Font(name='Calibri', size=14, bold=True, color='2D3436')
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{row}'].fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        ws.row_dimensions[row].height = 22
        
        row += 1
        
        # Create table headers
        headers = ['Metric', 'Value', 'Comparison', 'Analysis Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        ws.row_dimensions[row].height = 20
        
        row += 1
        
        # Enhanced metrics with more context
        metrics = [
            ("NPV Advantage", summary['npv_difference'], 
             f"${abs(summary['npv_difference']):,.0f} {'(Ownership)' if summary['npv_difference'] > 0 else '(Rental)'}", 
             "Net Present Value difference"),
            ("Ownership NPV", summary['ownership_npv'], 
             f"${summary['ownership_npv']:,.0f}", 
             "Total ownership scenario value"),
            ("Rental NPV", summary['rental_npv'], 
             f"${summary['rental_npv']:,.0f}", 
             "Total rental scenario value"),
            ("Initial Investment", summary['ownership_initial_investment'], 
             f"${summary['ownership_initial_investment']:,.0f}", 
             "Required upfront capital"),
            ("Analysis Period", f"{summary['analysis_period']} years", 
             f"{summary['analysis_period']} years", 
             "Investment time horizon"),
            ("Cost of Capital", f"{summary['cost_of_capital']:.1f}%", 
             f"{summary['cost_of_capital']:.1f}% annual discount rate", 
             "Required rate of return")
        ]
        
        for metric_name, metric_value, comparison, notes in metrics:
            # Metric name
            ws[f'A{row}'] = metric_name
            ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True, color='2D3436')
            ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'A{row}'].border = thin_border
            
            # Value
            ws[f'B{row}'] = metric_value
            ws[f'B{row}'].font = Font(name='Calibri', size=10, color='2D3436')
            ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'B{row}'].border = thin_border
            
            # Comparison
            ws[f'C{row}'] = comparison
            ws[f'C{row}'].font = Font(name='Calibri', size=10, color='2D3436')
            ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'C{row}'].border = thin_border
            
            # Notes
            ws[f'D{row}'] = notes
            ws[f'D{row}'].font = Font(name='Calibri', size=9, italic=True, color='636E72')
            ws[f'D{row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'D{row}'].border = thin_border
            
            # Format currency values
            if isinstance(metric_value, (int, float)) and abs(metric_value) > 1000:
                ws[f'B{row}'].number_format = '$#,##0'
            
            # Color code positive/negative values
            if isinstance(metric_value, (int, float)):
                if metric_value > 0 and 'NPV' in metric_name:
                    ws[f'B{row}'].font = Font(name='Calibri', size=10, color='00B894', bold=True)
            
            ws.row_dimensions[row].height = 18
            row += 1
    
    async def _create_cash_flows_sheet(
        self,
        ws: Worksheet,
        excel_data: Dict[str, Any],
        config: Dict[str, Any]
    ) -> None:
        """Create cash flows analysis worksheet"""
        cash_flows = excel_data['formatted_tables']['cash_flows']
        
        # Enhanced title with professional styling
        ws['A1'] = "CASH FLOW ANALYSIS & PROJECTIONS"
        ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='FF6B6B')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 25
        
        # Add subtitle with analysis context
        ws['A2'] = f"Multi-Scenario Financial Analysis - {datetime.now().strftime('%B %Y')}"
        ws['A2'].font = Font(name='Calibri', size=11, italic=True, color='636E72')
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A2:F2')
        ws.row_dimensions[2].height = 18
        
        current_row = 3
        
        # Ownership cash flows
        rows_used = await self._insert_data_table(ws, cash_flows['ownership_table'], start_row=current_row, title="Ownership Cash Flows")
        current_row += rows_used + 2  # Add spacing
        
        # Rental cash flows  
        rows_used = await self._insert_data_table(ws, cash_flows['rental_table'], start_row=current_row, title="Rental Cash Flows")
        current_row += rows_used + 2  # Add spacing
        
        # Comparison table
        await self._insert_data_table(ws, cash_flows['comparison_table'], start_row=current_row, title="Side-by-Side Comparison")
    
    async def _create_charts_sheet(
        self,
        ws: Worksheet,
        excel_data: Dict[str, Any],
        config: Dict[str, Any]
    ) -> None:
        """Create charts and visualizations worksheet"""
        # Render charts if not already done
        chart_images = excel_data.get('chart_images', {})
        if not chart_images:
            logger.info("Rendering charts for Excel embedding")
            chart_images = await self.render_charts(excel_data, resolution=300)
            excel_data['chart_images'] = chart_images
        
        ws['A1'] = "Charts & Visualizations"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Enhanced chart embedding with professional layout
        row = 3
        col = 1
        charts_per_page = 2  # Two charts per page for better organization
        charts_added = 0
        
        for chart_name, image_path in chart_images.items():
            if Path(image_path).exists():
                try:
                    img = Image(str(image_path))
                    
                    # Professional chart sizing for Excel
                    img.width = 650  # Optimized width for readability
                    img.height = 400  # Optimal height for screen viewing
                    
                    # Add professional chart title with styling
                    chart_title = chart_name.replace('_', ' ').title().replace('Npv', 'NPV')
                    title_cell = ws[f'{get_column_letter(col)}{row}']
                    title_cell.value = chart_title
                    title_cell.font = Font(name='Calibri', size=13, bold=True, color='2D3436')
                    title_cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Add subtle background for title
                    title_cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                    ws.merge_cells(f'{get_column_letter(col)}{row}:{get_column_letter(col+8)}{row}')
                    ws.row_dimensions[row].height = 22
                    
                    # Add image with proper positioning
                    ws.add_image(img, f'{get_column_letter(col)}{row + 1}')
                    
                    # Add chart description/context
                    context_row = row + 25  # Below the chart
                    context_cell = ws[f'{get_column_letter(col)}{context_row}']
                    
                    # Create contextual descriptions
                    descriptions = {
                        'NPV Comparison': 'Comparative analysis of Net Present Value between ownership and rental scenarios',
                        'Annual Cash Flows': 'Year-over-year cash flow projections for both investment scenarios', 
                        'Cumulative Cash Flows': 'Cumulative cash position over the analysis period',
                        'Financial Metrics': 'Key financial performance indicators and ratios',
                        'Sensitivity Analysis': 'Impact of key variables on investment outcomes'
                    }
                    
                    description = descriptions.get(chart_title, f'Analysis chart: {chart_title}')
                    context_cell.value = description
                    context_cell.font = Font(name='Calibri', size=9, italic=True, color='636E72')
                    context_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws.merge_cells(f'{get_column_letter(col)}{context_row}:{get_column_letter(col+8)}{context_row}')
                    ws.row_dimensions[context_row].height = 16
                    
                    row += 32  # Space for next chart (title + image + description + spacing)
                    charts_added += 1
                    
                except Exception as e:
                    logger.warning(f"Failed to embed chart {chart_name}: {str(e)}")
                    # Add a styled placeholder instead
                    placeholder_cell = ws[f'A{row}']
                    placeholder_cell.value = f"Chart: {chart_name} (Unable to display)"
                    placeholder_cell.font = Font(name='Calibri', size=10, italic=True, color='E74C3C')
                    placeholder_cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                    row += 3
            else:
                logger.warning(f"Chart image not found: {image_path}")
                # Styled placeholder for missing images
                missing_cell = ws[f'A{row}']
                missing_cell.value = f"Chart: {chart_name} (File not found)"
                missing_cell.font = Font(name='Calibri', size=10, italic=True, color='E67E22')
                missing_cell.fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
                row += 3
    
    async def _create_calculations_sheet(
        self,
        ws: Worksheet,
        excel_data: Dict[str, Any],
        config: Dict[str, Any]
    ) -> None:
        """Create detailed calculations worksheet"""
        calculations = excel_data['formatted_tables']['calculations']
        
        # Enhanced calculations title
        ws['A1'] = "DETAILED FINANCIAL CALCULATIONS"
        ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='FF6B6B')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 25
        
        # Add calculation methodology note
        ws['A2'] = "NPV Analysis, Mortgage Calculations, Tax Benefits & Terminal Value Computations"
        ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='636E72')
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A2:E2')
        ws.row_dimensions[2].height = 16
        
        current_row = 3
        
        # Insert each calculation section
        for calc_name, calc_data in calculations.items():
            title = calc_name.replace('_', ' ').title()
            rows_used = await self._insert_data_table(ws, calc_data, start_row=current_row, title=title)
            current_row += rows_used + 3  # Add spacing between sections
    
    async def _create_assumptions_sheet(
        self,
        ws: Worksheet,
        excel_data: Dict[str, Any],
        config: Dict[str, Any]
    ) -> None:
        """Create input assumptions worksheet"""
        assumptions = excel_data['formatted_tables']['assumptions']
        
        # Enhanced assumptions title
        ws['A1'] = "INPUT ASSUMPTIONS & MODEL PARAMETERS"
        ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='FF6B6B')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 25
        
        # Add assumptions context
        ws['A2'] = "Key Variables & Assumptions Used in Financial Modeling"
        ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='636E72')
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A2:E2')
        ws.row_dimensions[2].height = 16
        
        await self._insert_data_table(ws, assumptions, start_row=3, title="All Input Parameters")
    
    async def _insert_data_table(
        self,
        ws: Worksheet,
        table_data: Dict[str, Any],
        start_row: int,
        title: str
    ) -> int:
        """Insert formatted data table into worksheet"""
        
        # Section title
        ws[f'A{start_row}'] = title
        ws[f'A{start_row}'].font = Font(size=14, bold=True)
        
        # Apply professional table formatting using the formatter
        self.formatter.apply_table_formatting(ws, table_data, start_row + 2, 1)
        
        return len(table_data.get('data', [])) + 3  # Return the number of rows used
    
    async def _apply_workbook_styling(self, template_config: Dict[str, Any]) -> None:
        """Apply enhanced professional styling to entire workbook"""
        
        # Define professional color scheme consistent with PDF
        colors = {
            'primary': 'FF6B6B',
            'secondary': '74B9FF', 
            'success': '00B894',
            'warning': 'FDCB6E',
            'light': 'F8F9FA',
            'dark': '2D3436',
            'muted': '636E72',
            'white': 'FFFFFF',
            'border': 'D3D3D3'
        }
        
        for ws in self.workbook.worksheets:
            # Enhanced column width calculation with professional standards
            column_widths = {
                'A': 25,  # Metric/Category names
                'B': 18,  # Values
                'C': 20,  # Comparison/Analysis
                'D': 35,  # Notes/Descriptions
                'E': 25,  # Additional data (increased for better text visibility)
                'F': 15   # Additional data
            }
            
            # Apply optimized column widths with special handling for assumptions sheet
            for col_letter, width in column_widths.items():
                # Special handling for Input Assumptions sheet column E
                if ws.title == 'Input Assumptions' and col_letter == 'E':
                    ws.column_dimensions[col_letter].width = 40  # Extra wide for descriptions
                else:
                    ws.column_dimensions[col_letter].width = width
            
            # Auto-adjust remaining columns intelligently
            for column in ws.columns:
                column_letter = get_column_letter(column[0].column)
                
                if column_letter not in column_widths:
                    max_length = 0
                    for cell in column:
                        if cell.value is not None:
                            try:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                            except:
                                pass
                    
                    # Set intelligent width with bounds
                    adjusted_width = min(max(max_length + 3, 12), 40)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Apply enhanced cell formatting
            thin_border = Border(
                left=Side(style='thin', color=colors['border']),
                right=Side(style='thin', color=colors['border']),
                top=Side(style='thin', color=colors['border']),
                bottom=Side(style='thin', color=colors['border'])
            )
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell.border = thin_border
                        cell.alignment = Alignment(
                            vertical='center',
                            horizontal='left' if cell.column == 1 else 'center',
                            wrap_text=True
                        )
                        
                        # Apply consistent font styling
                        if cell.font.name != 'Calibri':  # Don't override already styled cells
                            cell.font = Font(
                                name='Calibri',
                                size=cell.font.size if cell.font.size else 10,
                                bold=cell.font.bold,
                                italic=cell.font.italic,
                                color=cell.font.color if cell.font.color else colors['dark']
                            )
            
            # Set professional row heights
            for row_num in range(1, ws.max_row + 1):
                if ws.row_dimensions[row_num].height is None:
                    ws.row_dimensions[row_num].height = 16
            
            # Apply worksheet-level formatting
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToHeight = False
            ws.page_setup.fitToWidth = 1
            
            # Add professional gridlines
            ws.sheet_view.showGridLines = True
    
    async def validate_data(self, export_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Validate export data for Excel generation
        
        Args:
            export_data: Export data package to validate
            
        Returns:
            Validation results with errors and warnings
        """
        validation_results = {
            'is_valid': True,
            'errors': [],
            'warnings': []
        }
        
        # Check required data keys (flexible for session_data/inputs)
        required_keys = ['analysis_results', 'ownership_flows', 'rental_flows']
        for key in required_keys:
            if key not in export_data:
                validation_results['errors'].append(f"Missing required key: {key}")
                validation_results['is_valid'] = False
        
        # Check for session_data or inputs (either is acceptable)
        if 'session_data' not in export_data and 'inputs' not in export_data:
            validation_results['errors'].append("Missing required key: 'session_data' or 'inputs'")
            validation_results['is_valid'] = False
        
        # Validate analysis results
        if 'analysis_results' in export_data:
            analysis = export_data['analysis_results']
            required_analysis_keys = ['ownership_npv', 'rental_npv', 'npv_difference', 'recommendation']
            for key in required_analysis_keys:
                if key not in analysis:
                    validation_results['warnings'].append(f"Missing analysis result: {key}")
        
        # Validate cash flows (handle both list and dict formats)
        if 'ownership_flows' in export_data and 'rental_flows' in export_data:
            ownership_flows = export_data['ownership_flows']
            rental_flows = export_data['rental_flows']
            
            # Handle list format (direct list of cash flow dicts)
            if isinstance(ownership_flows, list):
                if not ownership_flows:
                    validation_results['warnings'].append("No ownership cash flows data")
            # Handle dict format (with annual_cash_flows key)
            elif isinstance(ownership_flows, dict):
                if not ownership_flows.get('annual_cash_flows'):
                    validation_results['warnings'].append("No ownership cash flows data")
            else:
                validation_results['errors'].append("Ownership cash flows must be a list or dictionary")
                validation_results['is_valid'] = False
            
            # Same for rental flows
            if isinstance(rental_flows, list):
                if not rental_flows:
                    validation_results['warnings'].append("No rental cash flows data")
            elif isinstance(rental_flows, dict):
                if not rental_flows.get('annual_cash_flows'):
                    validation_results['warnings'].append("No rental cash flows data")
            else:
                validation_results['errors'].append("Rental cash flows must be a list or dictionary")
                validation_results['is_valid'] = False
        
        return validation_results
    
    def _normalize_cash_flows(self, cash_flows: Any) -> List[Dict[str, Any]]:
        """
        Normalize cash flows to a consistent format
        
        Args:
            cash_flows: Either a list of dicts or a dict with 'annual_cash_flows' key
            
        Returns:
            List of dictionaries with cash flow data
        """
        if isinstance(cash_flows, list):
            # Already in the expected format
            return cash_flows
        elif isinstance(cash_flows, dict):
            if 'annual_cash_flows' in cash_flows:
                annual_flows = cash_flows['annual_cash_flows']
                if isinstance(annual_flows, list):
                    # Convert simple array to list of dicts if needed
                    if annual_flows and not isinstance(annual_flows[0], dict):
                        return [
                            {'year': i+1, 'net_cash_flow': flow}
                            for i, flow in enumerate(annual_flows)
                        ]
                    else:
                        return annual_flows
            # If dict doesn't have annual_cash_flows, try to extract what we can
            return []
        else:
            return []
    
    def cleanup(self, defer_seconds: int = 0) -> None:
        """Clean up temporary resources
        
        Args:
            defer_seconds: Defer cleanup for this many seconds (useful for file access)
        """
        try:
            if self.workbook:
                self.workbook.close()
            
            if defer_seconds > 0:
                # Schedule cleanup for later
                import threading
                import time
                
                def deferred_cleanup():
                    time.sleep(defer_seconds)
                    self._do_cleanup()
                
                cleanup_thread = threading.Thread(target=deferred_cleanup, daemon=True)
                cleanup_thread.start()
                logger.debug(f"Scheduled cleanup of {self.temp_dir} in {defer_seconds} seconds")
            else:
                self._do_cleanup()
                    
        except Exception as e:
            logger.warning(f"Cleanup error: {e}")
    
    def _do_cleanup(self):
        """Actually perform the cleanup"""
        # Clean up temporary directory if it exists
        if self.temp_dir.exists():
            import shutil
            try:
                shutil.rmtree(self.temp_dir)
                logger.debug(f"Cleaned up temporary directory: {self.temp_dir}")
            except Exception as e:
                logger.warning(f"Could not clean up temp directory: {e}")
    
    def __del__(self):
        """Cleanup on destruction - but defer to allow file access"""
        self.cleanup(defer_seconds=5)