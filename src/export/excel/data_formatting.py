"""
Excel Data Formatting
Professional financial data formatting for Excel export

This module handles:
- Currency and percentage formatting for financial data
- Data table creation with proper headers and alignment
- Conditional formatting for positive/negative values
- Professional styling with corporate color schemes
- Number formatting for different data types (currency, percentage, date)

Repository: https://github.com/LT-aitools/rent-vs-buy-decision-tool
"""

import logging
from typing import Dict, List, Any, Optional, Tuple, Union
from datetime import datetime, date
import pandas as pd

from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class ExcelFormatter:
    """
    Professional Excel formatter for real estate financial data
    
    Handles formatting of financial data, tables, and styling
    for professional Excel workbooks.
    """
    
    # Enhanced professional color scheme consistent with PDF exports
    COLORS = {
        'primary': 'FF6B6B',        # Professional coral red
        'secondary': '74B9FF',      # Professional blue
        'success': '00B894',        # Strong success green
        'warning': 'FDCB6E',        # Warm warning yellow
        'danger': 'E17055',         # Muted danger red
        'info': 'A29BFE',           # Professional purple
        'light': 'F8F9FA',          # Clean light background
        'dark': '2D3436',           # Charcoal text
        'muted': '636E72',          # Professional muted text
        'accent': '74B9FF',         # Blue accent
        'neutral': 'DDD6D6',        # Neutral border
        'white': 'FFFFFF',          # White
        'black': '000000'           # Black
    }
    
    def __init__(self):
        """Initialize Excel formatter with enhanced professional styles"""
        self.currency_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
        self.currency_detailed_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        self.percentage_format = '0.00%_);[Red](0.00%)'
        self.number_format = '_* #,##0_);_* (#,##0);_* "-"_);_(@_)'
        self.date_format = 'mm/dd/yyyy'
        self.accounting_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
        
        logger.info("ExcelFormatter initialized with professional styling")
    
    async def format_cash_flow_table(
        self, 
        cash_flows: Union[Dict[str, Any], List[Dict[str, Any]]], 
        flow_type: str
    ) -> Dict[str, Any]:
        """
        Format cash flow data for Excel table display
        
        Args:
            cash_flows: Cash flow data from NPV analysis (list or dict format)
            flow_type: "ownership" or "rental"
            
        Returns:
            Formatted table data with headers and styled rows
        """
        logger.info(f"Formatting {flow_type} cash flow table")
        
        # Extract cash flow arrays - handle both list and dict formats
        annual_cash_flows = self._extract_cash_flows(cash_flows)
        years = list(range(len(annual_cash_flows)))
        
        # Create headers
        headers = ['Year', 'Annual Cash Flow', 'Cumulative Cash Flow', 'NPV Factor', 'Present Value']
        
        # Calculate cumulative and present values
        cumulative_flows = []
        npv_factors = []
        present_values = []
        
        # Extract cost of capital - handle both dict and list formats
        if isinstance(cash_flows, dict):
            cost_of_capital = cash_flows.get('cost_of_capital', 8.0) / 100
        else:
            # For list format, use default cost of capital
            cost_of_capital = 8.0 / 100
        
        cumulative = 0
        for i, annual_flow in enumerate(annual_cash_flows):
            cumulative += annual_flow
            cumulative_flows.append(cumulative)
            
            # NPV discount factor
            factor = 1 / ((1 + cost_of_capital) ** (i + 1))
            npv_factors.append(factor)
            
            # Present value of this year's cash flow
            pv = annual_flow * factor
            present_values.append(pv)
        
        # Create data rows
        data_rows = []
        for i in range(len(annual_cash_flows)):
            row = [
                years[i] + 1,  # Year (1-based)
                annual_cash_flows[i],
                cumulative_flows[i],
                npv_factors[i],
                present_values[i]
            ]
            data_rows.append(row)
        
        # Add totals row
        total_pv = sum(present_values)
        totals_row = [
            'TOTAL',
            sum(annual_cash_flows),
            cumulative_flows[-1] if cumulative_flows else 0,
            '',  # No total for NPV factors
            total_pv
        ]
        data_rows.append(totals_row)
        
        return {
            'headers': headers,
            'data': data_rows,
            'table_type': f'{flow_type}_cash_flows',
            'formatting_rules': {
                'currency_columns': [1, 2, 4],  # Annual, Cumulative, Present Value
                'percentage_columns': [],
                'total_row_index': len(data_rows) - 1,
                'conditional_formatting': {
                    'positive_cash_flow': {'column': 1, 'color': self.COLORS['success']},
                    'negative_cash_flow': {'column': 1, 'color': self.COLORS['danger']}
                }
            }
        }
    
    async def create_comparison_table(
        self, 
        ownership_flows: Union[Dict[str, Any], List[Dict[str, Any]]], 
        rental_flows: Union[Dict[str, Any], List[Dict[str, Any]]]
    ) -> Dict[str, Any]:
        """
        Create side-by-side comparison table of ownership vs rental cash flows
        
        Args:
            ownership_flows: Ownership cash flow data (list or dict format)
            rental_flows: Rental cash flow data (list or dict format)
            
        Returns:
            Formatted comparison table
        """
        logger.info("Creating ownership vs rental comparison table")
        
        ownership_annual = self._extract_cash_flows(ownership_flows)
        rental_annual = self._extract_cash_flows(rental_flows)
        
        # Ensure both arrays are same length
        max_years = max(len(ownership_annual), len(rental_annual))
        ownership_annual.extend([0] * (max_years - len(ownership_annual)))
        rental_annual.extend([0] * (max_years - len(rental_annual)))
        
        headers = [
            'Year', 
            'Ownership Cash Flow', 
            'Rental Cash Flow', 
            'Annual Difference', 
            'Cumulative Difference'
        ]
        
        data_rows = []
        cumulative_difference = 0
        
        for i in range(max_years):
            annual_diff = ownership_annual[i] - rental_annual[i]
            cumulative_difference += annual_diff
            
            row = [
                i + 1,  # Year (1-based)
                ownership_annual[i],
                rental_annual[i],
                annual_diff,
                cumulative_difference
            ]
            data_rows.append(row)
        
        # Add summary row
        ownership_total = sum(ownership_annual)
        rental_total = sum(rental_annual)
        total_difference = ownership_total - rental_total
        
        summary_row = [
            'TOTAL',
            ownership_total,
            rental_total,
            total_difference,
            cumulative_difference
        ]
        data_rows.append(summary_row)
        
        return {
            'headers': headers,
            'data': data_rows,
            'table_type': 'comparison',
            'formatting_rules': {
                'currency_columns': [1, 2, 3, 4],  # All cash flow columns
                'percentage_columns': [],
                'total_row_index': len(data_rows) - 1,
                'conditional_formatting': {
                    'ownership_advantage': {'column': 3, 'condition': '> 0', 'color': self.COLORS['success']},
                    'rental_advantage': {'column': 3, 'condition': '< 0', 'color': self.COLORS['danger']}
                }
            }
        }
    
    async def format_npv_calculations(self, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """
        Format NPV calculation details for Excel display
        
        Args:
            analysis_results: Complete NPV analysis results
            
        Returns:
            Formatted NPV calculations table
        """
        logger.info("Formatting NPV calculations")
        
        headers = ['Metric', 'Ownership', 'Rental', 'Difference']
        
        ownership_npv = analysis_results.get('ownership_npv', 0)
        rental_npv = analysis_results.get('rental_npv', 0)
        npv_difference = analysis_results.get('npv_difference', 0)
        
        ownership_investment = analysis_results.get('ownership_initial_investment', 0)
        rental_investment = analysis_results.get('rental_initial_investment', 0)
        investment_difference = ownership_investment - rental_investment
        
        ownership_irr = analysis_results.get('ownership_irr', 0)
        rental_irr = analysis_results.get('rental_irr', 0)
        irr_difference = ownership_irr - rental_irr
        
        data_rows = [
            ['Net Present Value', ownership_npv, rental_npv, npv_difference],
            ['Initial Investment', ownership_investment, rental_investment, investment_difference],
            ['Internal Rate of Return', ownership_irr, rental_irr, irr_difference],
            ['Recommendation', analysis_results.get('recommendation', 'UNKNOWN'), '', ''],
            ['Confidence Level', analysis_results.get('confidence', 'Low'), '', ''],
            ['Analysis Period', f"{analysis_results.get('analysis_period', 25)} years", '', ''],
            ['Cost of Capital', f"{analysis_results.get('cost_of_capital', 8)}%", '', '']
        ]
        
        return {
            'headers': headers,
            'data': data_rows,
            'table_type': 'npv_calculations',
            'formatting_rules': {
                'currency_columns': [1, 2, 3],  # NPV and investment columns
                'percentage_columns': [],  # IRR is handled specially
                'special_formatting': {
                    'irr_rows': [2],  # Format as percentage
                    'recommendation_row': 3,
                    'confidence_row': 4
                }
            }
        }
    
    async def format_mortgage_schedule(self, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """
        Format mortgage amortization schedule for Excel
        
        Args:
            analysis_results: Analysis results containing mortgage details
            
        Returns:
            Formatted mortgage schedule table
        """
        logger.info("Formatting mortgage schedule")
        
        # Extract mortgage parameters
        purchase_price = analysis_results.get('purchase_price', 500000)
        down_payment_pct = analysis_results.get('down_payment_percent', 30) / 100
        loan_amount = purchase_price * (1 - down_payment_pct)
        interest_rate = analysis_results.get('interest_rate', 5) / 100
        loan_term = analysis_results.get('loan_term', 20)
        
        # Calculate monthly payment
        monthly_rate = interest_rate / 12
        num_payments = loan_term * 12
        
        if monthly_rate > 0:
            monthly_payment = loan_amount * (monthly_rate * (1 + monthly_rate)**num_payments) / ((1 + monthly_rate)**num_payments - 1)
        else:
            monthly_payment = loan_amount / num_payments
        
        headers = [
            'Payment', 'Payment Amount', 'Principal', 'Interest', 
            'Remaining Balance', 'Cumulative Principal', 'Cumulative Interest'
        ]
        
        data_rows = []
        remaining_balance = loan_amount
        cumulative_principal = 0
        cumulative_interest = 0
        
        # Show first 12 months and then annual summaries
        for payment_num in range(1, min(13, num_payments + 1)):
            interest_payment = remaining_balance * monthly_rate
            principal_payment = monthly_payment - interest_payment
            remaining_balance -= principal_payment
            cumulative_principal += principal_payment
            cumulative_interest += interest_payment
            
            row = [
                payment_num,
                monthly_payment,
                principal_payment,
                interest_payment,
                remaining_balance,
                cumulative_principal,
                cumulative_interest
            ]
            data_rows.append(row)
        
        # Add summary information
        data_rows.append(['', '', '', '', '', '', ''])  # Blank row
        data_rows.append([
            'LOAN SUMMARY',
            f'${monthly_payment:,.2f}/month',
            '',
            '',
            '',
            '',
            ''
        ])
        data_rows.append([
            'Total Interest',
            '',
            '',
            monthly_payment * num_payments - loan_amount,
            '',
            '',
            ''
        ])
        
        return {
            'headers': headers,
            'data': data_rows,
            'table_type': 'mortgage_schedule',
            'formatting_rules': {
                'currency_columns': [1, 2, 3, 4, 5, 6],
                'percentage_columns': [],
                'summary_start_row': len(data_rows) - 3
            }
        }
    
    async def format_tax_calculations(self, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """
        Format tax calculation details
        
        Args:
            analysis_results: Analysis results containing tax details
            
        Returns:
            Formatted tax calculations table
        """
        logger.info("Formatting tax calculations")
        
        headers = ['Tax Item', 'Annual Amount', 'Tax Deductible', 'Tax Benefit', 'Description']
        
        # Extract tax-related parameters
        corporate_tax_rate = analysis_results.get('corporate_tax_rate', 25) / 100
        purchase_price = analysis_results.get('purchase_price', 500000)
        property_tax_rate = analysis_results.get('property_tax_rate', 1.2) / 100
        annual_rent = analysis_results.get('current_annual_rent', 120000)
        
        # Calculate tax items
        property_tax = purchase_price * property_tax_rate
        property_tax_benefit = property_tax * corporate_tax_rate if analysis_results.get('property_tax_deductible', True) else 0
        
        interest_expense = analysis_results.get('annual_interest_expense', 0)  # Would come from mortgage calc
        interest_tax_benefit = interest_expense * corporate_tax_rate if analysis_results.get('interest_deductible', True) else 0
        
        rent_expense = annual_rent
        rent_tax_benefit = rent_expense * corporate_tax_rate if analysis_results.get('rent_deductible', True) else 0
        
        data_rows = [
            [
                'Property Tax',
                property_tax,
                'Yes' if analysis_results.get('property_tax_deductible', True) else 'No',
                property_tax_benefit,
                'Annual property tax on owned building'
            ],
            [
                'Mortgage Interest',
                interest_expense,
                'Yes' if analysis_results.get('interest_deductible', True) else 'No',
                interest_tax_benefit,
                'Deductible mortgage interest expense'
            ],
            [
                'Rental Payments',
                rent_expense,
                'Yes' if analysis_results.get('rent_deductible', True) else 'No',
                rent_tax_benefit,
                'Annual rent payments (business expense)'
            ],
            ['', '', '', '', ''],  # Blank row
            [
                'Tax Rate Applied',
                f"{corporate_tax_rate:.1%}",
                '',
                '',
                'Corporate tax rate for deductions'
            ]
        ]
        
        return {
            'headers': headers,
            'data': data_rows,
            'table_type': 'tax_calculations',
            'formatting_rules': {
                'currency_columns': [1, 3],  # Annual Amount and Tax Benefit
                'percentage_columns': [],
                'summary_start_row': len(data_rows) - 1
            }
        }
    
    async def format_terminal_value(self, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """
        Format terminal value calculations
        
        Args:
            analysis_results: Analysis results containing terminal value details
            
        Returns:
            Formatted terminal value table
        """
        logger.info("Formatting terminal value calculations")
        
        headers = ['Component', 'Value', 'Calculation Method', 'Notes']
        
        # Extract terminal value components
        purchase_price = analysis_results.get('purchase_price', 500000)
        land_value_pct = analysis_results.get('land_value_percent', 25) / 100
        market_appreciation = analysis_results.get('market_appreciation_rate', 3) / 100
        analysis_period = analysis_results.get('analysis_period', 25)
        
        # Calculate components
        land_value = purchase_price * land_value_pct
        building_value = purchase_price * (1 - land_value_pct)
        
        # Appreciated land value
        appreciated_land_value = land_value * ((1 + market_appreciation) ** analysis_period)
        
        # Depreciated building value (simplified)
        depreciation_period = analysis_results.get('depreciation_period', 39)
        remaining_building_life = max(0, depreciation_period - analysis_period)
        remaining_building_value = building_value * (remaining_building_life / depreciation_period) if depreciation_period > 0 else 0
        
        # Total terminal value
        total_terminal_value = appreciated_land_value + remaining_building_value
        
        data_rows = [
            [
                'Original Purchase Price',
                purchase_price,
                'Input Parameter',
                'Initial property acquisition cost'
            ],
            [
                'Land Value (at purchase)',
                land_value,
                f'{land_value_pct:.1%} of purchase price',
                'Non-depreciating land component'
            ],
            [
                'Building Value (at purchase)',
                building_value,
                f'{1-land_value_pct:.1%} of purchase price',
                'Depreciating building component'
            ],
            ['', '', '', ''],  # Blank row
            [
                'Appreciated Land Value',
                appreciated_land_value,
                f'{market_appreciation:.1%}/year for {analysis_period} years',
                'Land value with market appreciation'
            ],
            [
                'Remaining Building Value',
                remaining_building_value,
                f'{remaining_building_life}/{depreciation_period} years remaining',
                'Depreciated building value'
            ],
            ['', '', '', ''],  # Blank row
            [
                'TOTAL TERMINAL VALUE',
                total_terminal_value,
                'Land + Building values',
                'Expected property value at end of analysis'
            ]
        ]
        
        return {
            'headers': headers,
            'data': data_rows,
            'table_type': 'terminal_value',
            'formatting_rules': {
                'currency_columns': [1],  # Value column
                'percentage_columns': [],
                'total_row_index': len(data_rows) - 1,
                'highlight_rows': [len(data_rows) - 1]  # Highlight total row
            }
        }
    
    async def format_assumptions_table(self, session_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Format input assumptions for Excel display
        
        Args:
            session_data: Complete session state data
            
        Returns:
            Formatted assumptions table
        """
        logger.info("Formatting input assumptions table")
        
        headers = ['Category', 'Parameter', 'Value', 'Unit', 'Description']
        
        data_rows = []
        
        # Extract actual values from potentially nested structure
        inputs_data = session_data.get('inputs', session_data) if isinstance(session_data.get('inputs'), dict) else session_data
        
        # Property Information
        data_rows.extend([
            ['Property', 'Purchase Price', inputs_data.get('purchase_price', 500000), '$', 'Property acquisition cost'],
            ['Property', 'Total Size', inputs_data.get('total_property_size', 5000), 'sq ft', 'Total building area'],
            ['Property', 'Space Needed', inputs_data.get('current_space_needed', 4000), 'sq ft', 'Current space requirements'],
            ['Property', 'Property Type', inputs_data.get('property_type', 'Commercial'), '', 'Building classification'],
            ['Property', 'Location', inputs_data.get('location', ''), '', 'Property location'],
            ['', '', '', '', ''],  # Blank row
        ])
        
        # Financial Parameters
        data_rows.extend([
            ['Finance', 'Down Payment', inputs_data.get('down_payment_percent', 30), '%', 'Initial equity percentage'],
            ['Finance', 'Interest Rate', inputs_data.get('interest_rate', 5.0), '%', 'Mortgage interest rate'],
            ['Finance', 'Loan Term', inputs_data.get('loan_term', 20), 'years', 'Mortgage amortization period'],
            ['Finance', 'Cost of Capital', inputs_data.get('cost_of_capital', 8.0), '%', 'Discount rate for NPV'],
            ['Finance', 'Corporate Tax Rate', inputs_data.get('corporate_tax_rate', 25), '%', 'Tax rate for deductions'],
            ['', '', '', '', ''],  # Blank row
        ])
        
        # Rental Parameters
        data_rows.extend([
            ['Rental', 'Annual Rent', inputs_data.get('current_annual_rent', 120000), '$', 'Current rental cost'],
            ['Rental', 'Rent Increase Rate', inputs_data.get('rent_increase_rate', 3.0), '%/year', 'Annual rent escalation'],
            ['Rental', 'Security Deposit', inputs_data.get('security_deposit_months', 2), 'months', 'Upfront security deposit'],
            ['Rental', 'Moving Costs', inputs_data.get('moving_costs', 10000), '$', 'One-time moving expense'],
            ['', '', '', '', ''],  # Blank row
        ])
        
        # Operating Costs
        data_rows.extend([
            ['Operations', 'Property Tax Rate', inputs_data.get('property_tax_rate', 1.2), '%/year', 'Annual property tax rate'],
            ['Operations', 'Insurance Cost', inputs_data.get('insurance_cost', 5000), '$/year', 'Annual insurance premium'],
            ['Operations', 'Maintenance', inputs_data.get('annual_maintenance_percent', 2.0), '%/year', 'Annual maintenance as % of value'],
            ['Operations', 'CapEx Reserve', inputs_data.get('longterm_capex_reserve', 1.5), '%/year', 'Capital expenditure reserve'],
            ['', '', '', '', ''],  # Blank row
        ])
        
        # Analysis Parameters
        data_rows.extend([
            ['Analysis', 'Analysis Period', inputs_data.get('analysis_period', 25), 'years', 'Investment holding period'],
            ['Analysis', 'Inflation Rate', inputs_data.get('inflation_rate', 3.0), '%/year', 'General inflation assumption'],
            ['Analysis', 'Market Appreciation', inputs_data.get('market_appreciation_rate', 3.0), '%/year', 'Property value appreciation'],
            ['Analysis', 'Currency', inputs_data.get('currency', 'USD'), '', 'Reporting currency'],
            ['Analysis', 'Analysis Date', inputs_data.get('analysis_date', datetime.now().strftime('%m/%d/%Y')), '', 'Date of analysis']
        ])
        
        return {
            'headers': headers,
            'data': data_rows,
            'table_type': 'assumptions',
            'formatting_rules': {
                'currency_columns': [2],  # Value column (context-dependent)
                'percentage_columns': [],
                'category_grouping': True,
                'blank_row_indices': [5, 11, 16, 21]  # Indices of blank rows for spacing
            }
        }
    
    def apply_table_formatting(
        self, 
        ws: Worksheet, 
        table_data: Dict[str, Any], 
        start_row: int, 
        start_col: int = 1
    ) -> None:
        """
        Apply professional formatting to a data table in Excel
        
        Args:
            ws: Excel worksheet
            table_data: Formatted table data with formatting rules
            start_row: Starting row for the table
            start_col: Starting column for the table
        """
        headers = table_data.get('headers', [])
        data_rows = table_data.get('data', [])
        formatting_rules = table_data.get('formatting_rules', {})
        
        # Apply enhanced professional header formatting
        for col_idx, header in enumerate(headers):
            col_letter = get_column_letter(start_col + col_idx)
            cell = ws[f'{col_letter}{start_row}']
            cell.value = header
            cell.font = Font(name='Calibri', size=11, bold=True, color=self.COLORS['white'])
            cell.fill = PatternFill(start_color=self.COLORS['primary'], end_color=self.COLORS['primary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Professional border styling
            cell.border = Border(
                left=Side(style='thin', color=self.COLORS['neutral']),
                right=Side(style='thin', color=self.COLORS['neutral']),
                top=Side(style='medium', color=self.COLORS['primary']),
                bottom=Side(style='medium', color=self.COLORS['primary'])
            )
            
            # Set appropriate row height for headers
            ws.row_dimensions[start_row].height = 22
        
        # Apply enhanced data row formatting
        for row_idx, data_row in enumerate(data_rows):
            excel_row = start_row + row_idx + 1
            
            for col_idx, cell_value in enumerate(data_row):
                col_letter = get_column_letter(start_col + col_idx)
                cell = ws[f'{col_letter}{excel_row}']
                cell.value = cell_value
                
                # Enhanced font and alignment with special handling for description columns
                cell.font = Font(name='Calibri', size=10, color=self.COLORS['dark'])
                
                # Special alignment for description columns (column 4 in assumptions table)
                if col_idx == 4 and table_data.get('table_type') == 'assumptions':
                    cell.alignment = Alignment(
                        horizontal='left',
                        vertical='center',
                        wrap_text=True,
                        indent=1
                    )
                else:
                    cell.alignment = Alignment(
                        horizontal='left' if col_idx == 0 else 'center',
                        vertical='center',
                        wrap_text=True
                    )
                
                # Enhanced number formatting
                if col_idx in formatting_rules.get('currency_columns', []):
                    if isinstance(cell_value, (int, float)):
                        cell.number_format = self.accounting_format
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col_idx in formatting_rules.get('percentage_columns', []):
                    if isinstance(cell_value, (int, float)):
                        cell.number_format = self.percentage_format
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                elif isinstance(cell_value, (int, float)) and abs(cell_value) >= 1:
                    cell.number_format = self.number_format
                
                # Apply enhanced conditional formatting
                self._apply_conditional_formatting(cell, cell_value, formatting_rules.get('conditional_formatting', {}))
                
                # Professional borders with subtle styling
                cell.border = Border(
                    left=Side(style='thin', color=self.COLORS['neutral']),
                    right=Side(style='thin', color=self.COLORS['neutral']),
                    top=Side(style='hair', color=self.COLORS['neutral']),
                    bottom=Side(style='hair', color=self.COLORS['neutral'])
                )
                
                # Enhanced total/summary row highlighting
                if row_idx == formatting_rules.get('total_row_index', -1):
                    cell.font = Font(name='Calibri', size=10, bold=True, color=self.COLORS['primary'])
                    cell.fill = PatternFill(start_color=self.COLORS['light'], end_color=self.COLORS['light'], fill_type='solid')
                    cell.border = Border(
                        left=Side(style='thin', color=self.COLORS['primary']),
                        right=Side(style='thin', color=self.COLORS['primary']),
                        top=Side(style='medium', color=self.COLORS['primary']),
                        bottom=Side(style='medium', color=self.COLORS['primary'])
                    )
                
                # Set appropriate row height
                ws.row_dimensions[excel_row].height = 18
    
    def _apply_conditional_formatting(
        self, 
        cell, 
        cell_value: Any, 
        conditional_rules: Dict[str, Any]
    ) -> None:
        """Apply enhanced conditional formatting based on cell value and context"""
        
        if isinstance(cell_value, (int, float)):
            # Enhanced positive/negative value formatting
            if cell_value > 0:
                for rule_name, rule_config in conditional_rules.items():
                    if 'positive' in rule_name.lower() or 'advantage' in rule_name.lower():
                        cell.fill = PatternFill(
                            start_color=rule_config['color'], 
                            end_color=rule_config['color'], 
                            fill_type='solid'
                        )
                        # Enhanced contrast and readability
                        if rule_config['color'] in ['00B894', 'FF7675']:  # Dark backgrounds
                            cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
                        else:
                            cell.font = Font(name='Calibri', size=10, bold=True, color=self.COLORS['dark'])
                            
            elif cell_value < 0:
                for rule_name, rule_config in conditional_rules.items():
                    if 'negative' in rule_name.lower() or 'disadvantage' in rule_name.lower():
                        cell.fill = PatternFill(
                            start_color=rule_config['color'], 
                            end_color=rule_config['color'], 
                            fill_type='solid'
                        )
                        # Enhanced contrast for negative values
                        if rule_config['color'] in ['E17055', 'FF7675']:  # Red backgrounds
                            cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
                        else:
                            cell.font = Font(name='Calibri', size=10, bold=True, color=self.COLORS['danger'])
            
            # Add magnitude-based formatting for large values
            if abs(cell_value) >= 100000:  # Values >= 100k
                current_font = cell.font
                cell.font = Font(
                    name=current_font.name or 'Calibri',
                    size=current_font.size or 10,
                    bold=True,
                    color=current_font.color or self.COLORS['dark']
                )
        
        # Enhanced string-based conditional formatting
        elif isinstance(cell_value, str):
            value_upper = cell_value.upper().strip()
            
            # Recommendation formatting
            for rule_name, rule_config in conditional_rules.items():
                if 'recommendation' in rule_name.lower():
                    if 'BUY' in value_upper:
                        cell.fill = PatternFill(
                            start_color=self.COLORS['success'], 
                            end_color=self.COLORS['success'], 
                            fill_type='solid'
                        )
                        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif 'RENT' in value_upper:
                        cell.fill = PatternFill(
                            start_color=self.COLORS['warning'], 
                            end_color=self.COLORS['warning'], 
                            fill_type='solid'
                        )
                        cell.font = Font(name='Calibri', size=11, bold=True, color=self.COLORS['dark'])
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Status and confidence level formatting
            if value_upper in ['HIGH', 'STRONG', 'EXCELLENT']:
                cell.font = Font(name='Calibri', size=10, bold=True, color=self.COLORS['success'])
            elif value_upper in ['MEDIUM', 'MODERATE', 'GOOD']:
                cell.font = Font(name='Calibri', size=10, bold=True, color=self.COLORS['info'])
            elif value_upper in ['LOW', 'WEAK', 'POOR']:
                cell.font = Font(name='Calibri', size=10, bold=True, color=self.COLORS['danger'])
            
            # Priority and importance indicators
            if 'ADVANTAGE' in value_upper or 'BENEFIT' in value_upper:
                cell.font = Font(name='Calibri', size=10, bold=True, color=self.COLORS['primary'])
    
    def get_currency_format(self, detailed: bool = False) -> str:
        """Get appropriate currency format string"""
        return self.currency_detailed_format if detailed else self.currency_format
    
    def get_percentage_format(self) -> str:
        """Get percentage format string"""
        return self.percentage_format
    
    def get_date_format(self) -> str:
        """Get date format string"""
        return self.date_format
    
    def get_number_format(self) -> str:
        """Get general number format string"""
        return self.number_format
    
    def format_currency_value(self, value: Union[int, float], detailed: bool = False) -> str:
        """
        Format a numeric value as currency string
        
        Args:
            value: Numeric value to format
            detailed: Whether to show cents
            
        Returns:
            Formatted currency string
        """
        if detailed:
            return f"${value:,.2f}"
        else:
            return f"${value:,.0f}"
    
    def format_percentage_value(self, value: Union[int, float], precision: int = 1) -> str:
        """
        Format a numeric value as percentage string
        
        Args:
            value: Numeric value to format (as decimal, e.g., 0.05 for 5%)
            precision: Number of decimal places
            
        Returns:
            Formatted percentage string
        """
        return f"{value*100:.{precision}f}%"
    
    def create_header_style(self) -> NamedStyle:
        """Create enhanced named style for professional table headers"""
        header_style = NamedStyle(name="professional_header")
        header_style.font = Font(
            name='Calibri', 
            size=11, 
            bold=True, 
            color=self.COLORS['white']
        )
        header_style.fill = PatternFill(
            start_color=self.COLORS['primary'], 
            end_color=self.COLORS['primary'], 
            fill_type='solid'
        )
        header_style.alignment = Alignment(
            horizontal='center', 
            vertical='center', 
            wrap_text=True
        )
        header_style.border = Border(
            left=Side(style='thin', color=self.COLORS['neutral']),
            right=Side(style='thin', color=self.COLORS['neutral']),
            top=Side(style='medium', color=self.COLORS['primary']),
            bottom=Side(style='medium', color=self.COLORS['primary'])
        )
        return header_style
    
    def create_currency_style(self, positive_color: str = None, negative_color: str = None) -> NamedStyle:
        """Create enhanced named style for professional currency formatting"""
        currency_style = NamedStyle(name="professional_currency")
        currency_style.font = Font(name='Calibri', size=10, color=self.COLORS['dark'])
        currency_style.number_format = self.accounting_format
        currency_style.alignment = Alignment(
            horizontal='right', 
            vertical='center',
            indent=1
        )
        currency_style.border = Border(
            left=Side(style='thin', color=self.COLORS['neutral']),
            right=Side(style='thin', color=self.COLORS['neutral']),
            top=Side(style='hair', color=self.COLORS['neutral']),
            bottom=Side(style='hair', color=self.COLORS['neutral'])
        )
        return currency_style
    
    def _extract_cash_flows(self, flows_data: Union[Dict[str, Any], List[Dict[str, Any]]]) -> List[float]:
        """
        Extract cash flow values from either list or dict format
        
        Args:
            flows_data: Either a list of flow dicts or a dict with 'annual_cash_flows' key
            
        Returns:
            List of cash flow values
        """
        if isinstance(flows_data, list):
            # List of flow dictionaries - extract 'net_cash_flow' values
            return [
                flow.get('net_cash_flow', 0) if isinstance(flow, dict) else float(flow)
                for flow in flows_data
            ]
        elif isinstance(flows_data, dict):
            if 'annual_cash_flows' in flows_data:
                # Dictionary format with 'annual_cash_flows' key
                annual_flows = flows_data['annual_cash_flows']
                if isinstance(annual_flows, list):
                    return [float(flow) for flow in annual_flows]
            # If dict doesn't have expected structure, return empty list
            return []
        else:
            # Unsupported format
            return []