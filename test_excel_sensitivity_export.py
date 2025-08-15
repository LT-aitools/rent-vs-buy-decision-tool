#!/usr/bin/env python3
"""
Test Excel Export of 2D Sensitivity Analysis
"""

import sys
import os
import asyncio

# Add src directory to path
sys.path.append(os.path.join(os.path.dirname(__file__), 'src'))

async def test_excel_sensitivity_export():
    """Test that the Excel export includes all 2D sensitivity analysis tables"""
    
    print("🔧 Testing Excel Export of 2D Sensitivity Analysis")
    print("=" * 60)
    
    try:
        from export.excel.excel_generator import ExcelGenerator
        
        # Mock export data similar to what the real app would provide
        mock_export_data = {
            'analysis_results': {
                'ownership_npv': -150000,
                'rental_npv': -400000,
                'npv_difference': 250000,
                'recommendation': 'BUY',
                'confidence': 'High',
                'ownership_initial_investment': 150000,
                'rental_initial_investment': 5000,
                'analysis_period': 25,
                'cost_of_capital': 8.0
            },
            'ownership_flows': [
                {'year': 1, 'net_cash_flow': -45000, 'mortgage_payment': 35000, 'property_taxes': 6000, 'insurance': 5000, 'maintenance': 10000, 'tax_benefits': 11000},
                {'year': 2, 'net_cash_flow': -46000, 'mortgage_payment': 35000, 'property_taxes': 6200, 'insurance': 5100, 'maintenance': 10200, 'tax_benefits': 11300}
            ],
            'rental_flows': [
                {'year': 1, 'net_cash_flow': -18000, 'annual_rent': 24000, 'tax_benefits': 6000},
                {'year': 2, 'net_cash_flow': -18500, 'annual_rent': 24500, 'tax_benefits': 6125}
            ],
            'session_data': {
                'inputs': {
                    'purchase_price': 500000,
                    'current_annual_rent': 24000,
                    'down_payment_percent': 30.0,
                    'interest_rate': 5.0,
                    'loan_term': 20,
                    'transaction_costs_percent': 5.0,
                    'rent_increase_rate': 3.0,
                    'analysis_period': 25,
                    'cost_of_capital': 8.0,
                    'property_tax_rate': 1.2,
                    'property_tax_escalation_rate': 2.0,
                    'insurance_cost': 5000,
                    'annual_maintenance_percent': 2.0,
                    'property_management': 0,
                    'longterm_capex_reserve': 1.5,
                    'obsolescence_risk_factor': 0.5,
                    'inflation_rate': 3.0,
                    'land_value_percent': 25.0,
                    'market_appreciation_rate': 3.0,
                    'depreciation_period': 39,
                    'corporate_tax_rate': 25.0,
                    'interest_deductible': True,
                    'property_tax_deductible': True,
                    'rent_deductible': True,
                    'moving_costs': 0.0,
                    'space_improvement_cost': 0.0,
                    'future_expansion_year': 'Never',
                    'additional_space_needed': 0,
                    'current_space_needed': 0,
                    'ownership_property_size': 0,
                    'rental_property_size': 0,
                    'subletting_potential': False,
                    'subletting_rate': 0,
                    'subletting_space_sqm': 0,
                    'property_upgrade_cycle': 30
                }
            }
        }
        
        print("📊 Creating Excel generator...")
        generator = ExcelGenerator()
        
        print("📋 Validating export data...")
        validation_result = await generator.validate_data(mock_export_data)
        
        if not validation_result['is_valid']:
            print("❌ Validation failed:")
            for error in validation_result['errors']:
                print(f"  • {error}")
            return False
        
        if validation_result['warnings']:
            print("⚠️ Validation warnings:")
            for warning in validation_result['warnings']:
                print(f"  • {warning}")
        
        print("✅ Data validation passed")
        
        print("🔍 Preparing Excel data...")
        excel_data = await generator.prepare_data(mock_export_data)
        
        # Check if sensitivity data was prepared
        sensitivity_data = excel_data['formatted_tables'].get('sensitivity_analysis', {})
        
        if not sensitivity_data:
            print("❌ No sensitivity analysis data was generated")
            return False
        
        print(f"✅ Generated {len(sensitivity_data)} sensitivity tables:")
        for table_name in sensitivity_data.keys():
            print(f"  • {table_name}")
        
        # Verify we have all expected combinations
        from calculations.two_dimensional_sensitivity import get_available_sensitivity_metrics
        available_metrics = get_available_sensitivity_metrics()
        metric_keys = list(available_metrics.keys())
        
        expected_tables = []
        for y_metric in metric_keys:
            for x_metric in metric_keys:
                if x_metric != y_metric:
                    table_name = f"{available_metrics[y_metric]} vs {available_metrics[x_metric]}"
                    expected_tables.append(table_name)
        
        print(f"📈 Expected {len(expected_tables)} tables, got {len(sensitivity_data)}")
        
        missing_tables = []
        for expected in expected_tables:
            if expected not in sensitivity_data:
                missing_tables.append(expected)
        
        if missing_tables:
            print("❌ Missing sensitivity tables:")
            for missing in missing_tables:
                print(f"  • {missing}")
            return False
        
        print("✅ All expected sensitivity tables are present")
        
        print("📄 Generating Excel workbook...")
        output_path = await generator.generate_workbook(excel_data, template_type="detailed")
        
        print(f"✅ Excel workbook generated: {output_path}")
        print(f"   File exists: {output_path.exists()}")
        print(f"   File size: {output_path.stat().st_size if output_path.exists() else 0} bytes")
        
        # Test that the workbook can be opened and contains the sensitivity sheet
        if output_path.exists():
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(output_path)
                
                print("📊 Workbook contents:")
                for sheet_name in workbook.sheetnames:
                    print(f"  • {sheet_name}")
                
                if 'Sensitivity Analysis' in workbook.sheetnames:
                    print("✅ Sensitivity Analysis worksheet found in Excel file")
                    
                    # Check the content of the sensitivity sheet
                    sensitivity_sheet = workbook['Sensitivity Analysis']
                    
                    # Check for the title
                    title_cell = sensitivity_sheet['A1']
                    if title_cell.value and 'SENSITIVITY' in str(title_cell.value).upper():
                        print("✅ Sensitivity worksheet has proper title")
                    else:
                        print(f"❌ Unexpected title: {title_cell.value}")
                    
                    # Count non-empty cells to verify content
                    non_empty_cells = 0
                    for row in sensitivity_sheet.iter_rows():
                        for cell in row:
                            if cell.value is not None:
                                non_empty_cells += 1
                    
                    print(f"📊 Sensitivity worksheet has {non_empty_cells} populated cells")
                    
                    if non_empty_cells > 100:  # Should have lots of content with all tables
                        print("✅ Sensitivity worksheet appears to contain substantial data")
                    else:
                        print("⚠️ Sensitivity worksheet may not have enough content")
                    
                else:
                    print("❌ Sensitivity Analysis worksheet not found in Excel file")
                    return False
                
                workbook.close()
                
            except Exception as e:
                print(f"❌ Error reading Excel file: {e}")
                return False
        
        # Cleanup
        generator.cleanup()
        
        print("\n🎉 Excel sensitivity export test completed successfully!")
        return True
        
    except Exception as e:
        print(f"❌ Test failed with error: {e}")
        import traceback
        traceback.print_exc()
        return False

async def test_template_configuration():
    """Test that the template includes the sensitivity worksheet"""
    
    print("\n🔧 Testing Template Configuration")
    print("=" * 40)
    
    try:
        from export.excel.template_manager import ExcelTemplateManager
        
        manager = ExcelTemplateManager()
        
        # Get detailed template config
        detailed_config = await manager.get_template_config("detailed")
        
        print("📋 Detailed template worksheets:")
        worksheets = detailed_config.get('worksheets', [])
        
        for ws in worksheets:
            print(f"  • {ws['name']} (type: {ws['type']}, order: {ws['order']})")
        
        # Check if sensitivity worksheet is included
        sensitivity_sheets = [ws for ws in worksheets if ws['type'] == 'sensitivity']
        
        if sensitivity_sheets:
            print("✅ Sensitivity Analysis worksheet found in template")
            sensitivity_sheet = sensitivity_sheets[0]
            print(f"   Name: {sensitivity_sheet['name']}")
            print(f"   Order: {sensitivity_sheet['order']}")
            print(f"   Sections: {sensitivity_sheet['sections']}")
        else:
            print("❌ Sensitivity Analysis worksheet not found in template")
            return False
        
        return True
        
    except Exception as e:
        print(f"❌ Template test failed: {e}")
        return False

if __name__ == "__main__":
    print("🚀 Running Excel Sensitivity Export Tests")
    print()
    
    async def run_all_tests():
        success = True
        
        # Test template configuration
        if not await test_template_configuration():
            success = False
        
        # Test Excel export
        if not await test_excel_sensitivity_export():
            success = False
        
        print("\n" + "=" * 60)
        if success:
            print("🎉 ALL TESTS PASSED!")
            print("\n✅ Excel export includes comprehensive 2D sensitivity analysis")
            print("✅ All metric permutations are exported as separate tables")
            print("✅ Template configuration is correct")
            print("✅ Excel file can be generated and opened successfully")
        else:
            print("❌ SOME TESTS FAILED - Check errors above")
        
        print("=" * 60)
        return success
    
    # Run the async tests
    success = asyncio.run(run_all_tests())
    exit(0 if success else 1)